import fitz
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from typing import List, Tuple, Dict, Optional
import numpy as np
from dataclasses import dataclass
import os
import logging
from concurrent.futures import ThreadPoolExecutor
from time import time

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@dataclass
class TextBlock:
    """Represents a text block with its position and content."""
    x0: float
    y0: float
    x1: float
    y1: float
    text: str
    block_no: int
    page_num: int
    font_size: float
    font_name: str

@dataclass
class Cell:
    """Represents a table cell with potential spanning information."""
    content: str
    rowspan: int = 1
    colspan: int = 1
    x0: float = 0.0
    y0: float = 0.0
    x1: float = 0.0
    y1: float = 0.0

@dataclass
class Table:
    """Represents a detected table with its cells and position."""
    rows: List[List[Cell]]
    x0: float
    y0: float
    x1: float
    y1: float
    page_num: int
    confidence_score: float
    has_borders: bool
    header_row: Optional[List[str]] = None

class PDFTableExtractor:
    def __init__(self, tolerance: float = 3.0, min_rows: int = 2, 
                 min_cols: int = 2, max_cell_height: float = 50.0,
                 line_margin: float = 2.0):
        """
        Initialize the table extractor with configurable parameters.
        
        Args:
            tolerance: Tolerance for considering text blocks in the same row/column
            min_rows: Minimum number of rows required to consider a structure as table
            min_cols: Minimum number of columns required to consider a structure as table
            max_cell_height: Maximum height for single-line cells
            line_margin: Margin for detecting table lines/borders
        """
        self.tolerance = tolerance
        self.min_rows = min_rows
        self.min_cols = min_cols
        self.max_cell_height = max_cell_height
        self.line_margin = line_margin

    def _get_text_blocks(self, page: fitz.Page, page_num: int) -> List[TextBlock]:
        """Extract text blocks from a page with their positions and properties."""
        blocks = []
        
        # First, get all the lines (borders) in the page
        lines = page.get_drawings()
        line_coords = []
        for line in lines:
            if line["type"] == "l":  # Line
                line_coords.append((
                    min(line["rect"][0], line["rect"][2]),
                    min(line["rect"][1], line["rect"][3]),
                    max(line["rect"][0], line["rect"][2]),
                    max(line["rect"][1], line["rect"][3])
                ))

        # Get text blocks with their font information
        for block_no, block in enumerate(page.get_text("dict")["blocks"]):
            if "lines" not in block:
                continue
            
            block_text = []
            block_bbox = [float('inf'), float('inf'), float('-inf'), float('-inf')]
            font_sizes = []
            font_names = set()
            
            for line in block["lines"]:
                for span in line["spans"]:
                    text = span["text"].strip()
                    if text:
                        block_text.append(text)
                        # Update bounding box
                        block_bbox[0] = min(block_bbox[0], span["bbox"][0])
                        block_bbox[1] = min(block_bbox[1], span["bbox"][1])
                        block_bbox[2] = max(block_bbox[2], span["bbox"][2])
                        block_bbox[3] = max(block_bbox[3], span["bbox"][3])
                        font_sizes.append(span["size"])
                        font_names.add(span["font"])
            
            if block_text:
                # Check if block is surrounded by lines (table cell)
                is_cell = False
                for x0, y0, x1, y1 in line_coords:
                    if (abs(x0 - block_bbox[0]) <= self.line_margin or 
                        abs(x1 - block_bbox[2]) <= self.line_margin or 
                        abs(y0 - block_bbox[1]) <= self.line_margin or 
                        abs(y1 - block_bbox[3]) <= self.line_margin):
                        is_cell = True
                        break
                
                blocks.append(TextBlock(
                    x0=block_bbox[0],
                    y0=block_bbox[1],
                    x1=block_bbox[2],
                    y1=block_bbox[3],
                    text=" ".join(block_text),
                    block_no=block_no,
                    page_num=page_num,
                    font_size=np.mean(font_sizes),
                    font_name=list(font_names)[0] if font_names else ""
                ))
        
        return blocks

    def _identify_table_regions(self, blocks: List[TextBlock]) -> List[List[TextBlock]]:
        """Group text blocks into potential table regions using advanced detection."""
        if not blocks:
            return []

        # Sort blocks by y-coordinate
        blocks.sort(key=lambda b: b.y0)
        
        # Find common font characteristics for potential headers
        font_sizes = [b.font_size for b in blocks]
        common_font_size = np.median(font_sizes)
        
        # Group blocks by vertical position with adaptive tolerance
        table_rows = []
        current_row = [blocks[0]]
        last_y = blocks[0].y0
        
        for block in blocks[1:]:
            # Adjust tolerance based on font size and position
            local_tolerance = self.tolerance
            if block.font_size > common_font_size:
                local_tolerance *= 1.5  # More tolerance for headers
            
            if abs(block.y0 - last_y) <= local_tolerance:
                current_row.append(block)
            else:
                if current_row:
                    # Sort row by x-coordinate
                    current_row.sort(key=lambda b: b.x0)
                    table_rows.append(current_row)
                current_row = [block]
                last_y = block.y0
        
        if current_row:
            current_row.sort(key=lambda b: b.x0)
            table_rows.append(current_row)
        
        return table_rows

    def _detect_columns(self, table_rows: List[List[TextBlock]]) -> List[float]:
        """Detect column positions using advanced alignment analysis."""
        if not table_rows:
            return []

        # Collect all x-coordinates
        x_coords = []
        for row in table_rows:
            for block in row:
                x_coords.extend([block.x0, block.x1])
        
        # Use kernel density estimation to find column boundaries
        x_coords = np.array(x_coords)
        kde = np.histogram(x_coords, bins=min(len(x_coords), 50))
        peaks = []
        
        # Find peaks in the density
        for i in range(1, len(kde[0]) - 1):
            if kde[0][i] > kde[0][i-1] and kde[0][i] > kde[0][i+1]:
                peaks.append(kde[1][i])
        
        # Merge nearby peaks
        merged_peaks = []
        for peak in sorted(peaks):
            if not merged_peaks or abs(peak - merged_peaks[-1]) > self.tolerance:
                merged_peaks.append(peak)
        
        return merged_peaks

    def _detect_table_structure(self, page: fitz.Page) -> List[Tuple[float, float, float, float]]:
        """Detect table structure using lines and spacing analysis."""
        # Get all lines from the page
        drawings = page.get_drawings()
        horizontal_lines = []
        vertical_lines = []
        
        for drawing in drawings:
            if drawing["type"] == "l":  # Line
                x0, y0, x1, y1 = drawing["rect"]
                if abs(y1 - y0) <= self.line_margin:  # Horizontal line
                    horizontal_lines.append((min(x0, x1), y0, max(x0, x1), y1))
                elif abs(x1 - x0) <= self.line_margin:  # Vertical line
                    vertical_lines.append((x0, min(y0, y1), x1, max(y0, y1)))
        
        # Find table boundaries from intersecting lines
        table_regions = []
        for h1 in horizontal_lines:
            for h2 in horizontal_lines:
                if h1 != h2 and abs(h1[0] - h2[0]) <= self.tolerance:
                    # Found potential table top and bottom
                    y_top = min(h1[1], h2[1])
                    y_bottom = max(h1[1], h2[1])
                    
                    # Look for vertical lines within this range
                    matching_verticals = [
                        v for v in vertical_lines
                        if v[1] <= y_top + self.tolerance and v[3] >= y_bottom - self.tolerance
                    ]
                    
                    if len(matching_verticals) >= 2:
                        x_left = min(v[0] for v in matching_verticals)
                        x_right = max(v[0] for v in matching_verticals)
                        table_regions.append((x_left, y_top, x_right, y_bottom))
        
        return table_regions

    def _assign_cells_to_columns(self, row: List[TextBlock], columns: List[float]) -> List[Cell]:
        """Assign text blocks to columns with improved accuracy."""
        result = []
        used_blocks = set()
        
        for i, col_start in enumerate(columns):
            col_end = columns[i + 1] if i + 1 < len(columns) else float('inf')
            matching_blocks = []
            
            for block in row:
                if block not in used_blocks:
                    # Check if block belongs to this column
                    block_center = (block.x0 + block.x1) / 2
                    if col_start - self.tolerance <= block_center <= col_end + self.tolerance:
                        matching_blocks.append(block)
                        used_blocks.add(block)
            
            if matching_blocks:
                # Sort blocks vertically if multiple blocks in same column
                matching_blocks.sort(key=lambda b: b.y0)
                content = "\n".join(block.text for block in matching_blocks)
                result.append(Cell(
                    content=content,
                    x0=min(b.x0 for b in matching_blocks),
                    y0=min(b.y0 for b in matching_blocks),
                    x1=max(b.x1 for b in matching_blocks),
                    y1=max(b.y1 for b in matching_blocks)
                ))
            else:
                result.append(Cell(""))
        
        return result

    def extract_tables(self, pdf_path: str) -> List[Table]:
        """Extract tables from a PDF file with improved accuracy."""
        start_time = time()
        logger.info(f"Starting table extraction from {pdf_path}")
        
        doc = fitz.open(pdf_path)
        all_tables = []
        
        try:
            for page_num in range(len(doc)):
                page = doc[page_num]
                logger.debug(f"Processing page {page_num + 1}")
                
                # Detect table regions from lines
                table_regions = self._detect_table_structure(page)
                
                # Get all text blocks
                blocks = self._get_text_blocks(page, page_num)
                
                # Process each table region
                for region in table_regions:
                    # Filter blocks within this region
                    region_blocks = [
                        b for b in blocks
                        if (region[0] - self.tolerance <= b.x0 <= region[2] + self.tolerance and
                            region[1] - self.tolerance <= b.y0 <= region[3] + self.tolerance)
                    ]
                    
                    if len(region_blocks) >= self.min_rows * self.min_cols:
                        table_rows = self._identify_table_regions(region_blocks)
                        columns = self._detect_columns(table_rows)
                        
                        if len(table_rows) >= self.min_rows and len(columns) >= self.min_cols:
                            processed_rows = [
                                self._assign_cells_to_columns(row, columns)
                                for row in table_rows
                            ]
                            
                            table = Table(
                                rows=processed_rows,
                                x0=region[0],
                                y0=region[1],
                                x1=region[2],
                                y1=region[3],
                                page_num=page_num,
                                confidence_score=0.8,  # High confidence for bordered tables
                                has_borders=True,
                                header_row=[row[0].content for row in processed_rows[:1]] if processed_rows else None
                            )
                            all_tables.append(table)
                
                # Look for borderless tables in remaining blocks
                remaining_blocks = [
                    b for b in blocks
                    if not any(
                        region[0] - self.tolerance <= b.x0 <= region[2] + self.tolerance and
                        region[1] - self.tolerance <= b.y0 <= region[3] + self.tolerance
                        for region in table_regions
                    )
                ]
                
                if remaining_blocks:
                    table_rows = self._identify_table_regions(remaining_blocks)
                    if len(table_rows) >= self.min_rows:
                        columns = self._detect_columns(table_rows)
                        if len(columns) >= self.min_cols:
                            processed_rows = [
                                self._assign_cells_to_columns(row, columns)
                                for row in table_rows
                            ]
                            
                            # Calculate table boundaries
                            x0 = min(cell.x0 for row in processed_rows for cell in row if cell.content)
                            y0 = min(cell.y0 for row in processed_rows for cell in row if cell.content)
                            x1 = max(cell.x1 for row in processed_rows for cell in row if cell.content)
                            y1 = max(cell.y1 for row in processed_rows for cell in row if cell.content)
                            
                            table = Table(
                                rows=processed_rows,
                                x0=x0,
                                y0=y0,
                                x1=x1,
                                y1=y1,
                                page_num=page_num,
                                confidence_score=0.6,  # Lower confidence for borderless tables
                                has_borders=False,
                                header_row=[row[0].content for row in processed_rows[:1]] if processed_rows else None
                            )
                            all_tables.append(table)
            
            logger.info(f"Found {len(all_tables)} tables")
            
        finally:
            doc.close()
            
        end_time = time()
        logger.info(f"Table extraction completed in {end_time - start_time:.2f} seconds")
        return all_tables

    def save_to_excel(self, tables: List[Table], output_path: str):
        """Save extracted tables to Excel with improved formatting."""
        wb = Workbook()
        wb.remove(wb.active)
        
        # Define styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        center_aligned = Alignment(horizontal='center', vertical='center')
        
        for i, table in enumerate(tables):
            ws = wb.create_sheet(title=f'Table_{i+1}_Page_{table.page_num + 1}')
            
            # Write header row with special formatting if available
            if table.header_row:
                for col, header in enumerate(table.header_row, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.border = thin_border
                    cell.alignment = center_aligned
                    cell.font = cell.font.copy(bold=True)
            
            # Write data rows
            for row_idx, row in enumerate(table.rows, 1):
                for col_idx, cell in enumerate(row, 1):
                    excel_cell = ws.cell(row=row_idx, column=col_idx, value=cell.content)
                    
                    # Apply formatting
                    excel_cell.border = thin_border
                    excel_cell.alignment = center_aligned
                    
                    # Handle merged cells
                    if cell.rowspan > 1 or cell.colspan > 1:
                        ws.merge_cells(
                            start_row=row_idx,
                            start_column=col_idx,
                            end_row=row_idx + cell.rowspan - 1,
                            end_column=col_idx + cell.colspan - 1
                        )
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        wb.save(output_path)
        logger.info(f"Tables saved to {output_path}")

def process_pdf(pdf_path: str, output_dir: str, tolerance: float = 3.0):
    """
    Process a single PDF file and save its tables to Excel.
    
    Args:
        pdf_path: Path to the PDF file
        output_dir: Directory to save the output Excel file
        tolerance: Tolerance value for table detection
    """
    try:
        extractor = PDFTableExtractor(tolerance=tolerance)
        tables = extractor.extract_tables(pdf_path)
        
        if tables:
            output_path = os.path.join(
                output_dir,
                f"{os.path.splitext(os.path.basename(pdf_path))[0]}_tables.xlsx"
            )
            extractor.save_to_excel(tables, output_path)
            return True, f"Successfully processed {pdf_path}"
        else:
            return False, f"No tables found in {pdf_path}"
    except Exception as e:
        return False, f"Error processing {pdf_path}: {str(e)}"

def batch_process_pdfs(pdf_dir: str, output_dir: str, max_workers: int = 4):
    """
    Process multiple PDF files in parallel.
    
    Args:
        pdf_dir: Directory containing PDF files
        output_dir: Directory to save output Excel files
        max_workers: Maximum number of parallel workers
    """
    os.makedirs(output_dir, exist_ok=True)
    pdf_files = [f for f in os.listdir(pdf_dir) if f.lower().endswith('.pdf')]
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = []
        for pdf_file in pdf_files:
            pdf_path = os.path.join(pdf_dir, pdf_file)
            futures.append(
                executor.submit(process_pdf, pdf_path, output_dir)
            )
        
        for future in futures:
            success, message = future.result()
            if success:
                logger.info(message)
            else:
                logger.error(message)

def main():
    """Example usage of the table extractor."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Extract tables from PDF files')
    parser.add_argument('input', help='Input PDF file or directory')
    parser.add_argument('output', help='Output directory for Excel files')
    parser.add_argument('--tolerance', type=float, default=3.0,
                      help='Tolerance for table detection (default: 3.0)')
    parser.add_argument('--batch', action='store_true',
                      help='Process multiple PDFs in parallel')
    parser.add_argument('--workers', type=int, default=4,
                      help='Number of parallel workers for batch processing')
    
    args = parser.parse_args()
    
    try:
        if args.batch:
            if not os.path.isdir(args.input):
                raise ValueError("Input must be a directory when using batch processing")
            batch_process_pdfs(args.input, args.output, args.workers)
        else:
            if not os.path.isfile(args.input):
                raise ValueError("Input must be a file when not using batch processing")
            success, message = process_pdf(args.input, args.output, args.tolerance)
            if success:
                logger.info(message)
            else:
                logger.error(message)
    except Exception as e:
        logger.error(f"Error: {str(e)}")
        exit(1)

if __name__ == "__main__":
    main() 